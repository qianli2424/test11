# -*- coding=utf8 -*-
#@author:qianli
# 文件说明：
import openpyxl


def hide(*args):    #args前面加*，表示()中可以有0个参数，1个参数或多个参数
    file = openpyxl.load_workbook('test1.xlsx')
    sheet1 = file['Sheet']
    lastrow = sheet1.max_row
    i = 1
    for arg in args:  #args在函数中会被当成一个元素，这个元素中可以用for循环依次读取出来
        #也可以用len(args)获得这个元组的长度（个数），也可以用args[0]获取第1个元素的值
        sheet1.cell(row=lastrow+1,column=i).value=arg    #相当于把元组的所有值依次保存到excel中
        i = i+1
    file.save('test1.xlsx')
group = ['张三','李四','老王']
hide(*group)

#*args -> args 其实做了一个组合    args->*args 分离，解包