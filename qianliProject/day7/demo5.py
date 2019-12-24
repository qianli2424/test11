# -*- coding=utf8 -*-
#@author:qianli
# 文件说明：
import openpyxl
name = 'test3.xlsx'
try:
    f = openpyxl.load_workbook(name)
except Exception as ffe:
    print(ffe)
else:   #执行第7行就执行以下代码，执行第9行，呵呵，不好意思。
    sname = f.sheetnames[0]
    sheet1 = f[sname]
    sheet1.cell(1,1).values='python'
    f.save(name)
print('文件操作结束')