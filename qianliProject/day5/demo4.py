# -*- coding=utf8 -*-
#@author:qianli
# 文件说明：
import openpyxl

file = openpyxl.load_workbook('test.xlsx')
sheet1 = file['Sheet1']
maxrow = sheet1.max_row
maxcol = sheet1.max_column

users = []
#得到n行数据
for x in range(maxrow-1):
    # 得到一行数据
    user = []
    for i in range(maxcol-2):
        data = sheet1.cell(x+2, i+2).value
        user.append(data)
    #print(user)     #user是一行内容
    users.append(user)  #我是把一整行内容放到列表中
print(users)

name = input('请输入用户名')
password = input('请输入密码')
user = [name,password,'unlock']
if user in users:
    print('登录成功')
else:
    print('登录失败')