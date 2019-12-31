# -*- coding=utf8 -*-
#@author:qianli
# 文件说明：
import openpyxl

class excel:
    paths = 'C:\\program file\\python'
    def __init__(self,filename,sheetname):
        #在init方法中通过self.xx定义的xx属性叫做对象属性
        #我们创建了对象，可以使用对象名.xx来操作它（输出值，重新赋值）
        self.wb = openpyxl.load_workbook(filename)
        self.sheet = self.wb[sheetname]
        print(self.paths)

    def dictline(self,row=1,start=2,end=7):
        case = {}
        for i in range(start,end+1):
            key = self.sheet.cell(row=1,column=i).value
            value = self.sheet.cell(row=row,column=i).value
            case[key]=value     #添加一个元素
        return case
    def test(self):
        self.dictline()

if __name__ == '__main__':
    xls = excel('../testdata/case.xlsx','Sheet1')
    xls.wb = 'hello'
    print(xls.wb,xls.sheet)
    case = xls.dictline(2,2,6)
    print(case)
